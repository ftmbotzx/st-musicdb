import logging
import re
import time
import asyncio
from datetime import datetime
from pyrogram.client import Client
from pyrogram import filters
from pyrogram.types import Message
from pyrogram.errors import ChannelPrivate, ChatAdminRequired, UsernameNotOccupied, FloodWait
from bot.database import DatabaseManager
from bot.utils import extract_track_info, format_file_caption, get_file_metadata

logger = logging.getLogger(__name__)

# Initialize database manager
db = DatabaseManager()

# Global variable to track indexing process
indexing_process = {
    "active": False,
    "chat_id": None,
    "message_id": None,
    "total": 0,
    "processed": 0,
    "stop_requested": False
}

async def handle_media_message(client: Client, message: Message):
    """Handle incoming media messages and index them"""
    try:
        # Extract file metadata based on message type
        file_data = get_file_metadata(message)
        if not file_data:
            return
        
        # Check if file already exists in database to prevent duplicates
        try:
            existing_file = db.get_file_by_id(file_data["file_id"])
            if existing_file:
                logger.info(f"File already indexed, skipping: {file_data['file_id']}")
                return
                
            # Check if unique_id exists to prevent duplicates
            unique_id = file_data.get("file_unique_id")
            if unique_id:
                existing_unique = db.get_file_by_unique_id(unique_id)
                if existing_unique:
                    logger.info(f"File with unique_id already indexed, skipping: {unique_id}")
                    return
        except AttributeError as e:
            logger.warning(f"Database method not available: {e}, proceeding with indexing")
        
        # Extract track information from ALL text sources including entities
        message_text = message.text or message.caption or ""
        
        # CRITICAL: Check for \u2063 characters in the original message
        if '\u2063' in message_text:
            logger.info(f"FOUND \\u2063 in message: {repr(message_text)}")
            # Log each character for analysis
            for i, char in enumerate(message_text):
                if char == '\u2063':
                    logger.info(f"\\u2063 found at position {i}")
                elif ord(char) > 127:
                    logger.info(f"Special char at {i}: {repr(char)} (code: {ord(char)})")
        
        # Also check for URLs in message entities (clickable links) and CAPTION entities
        entity_urls = []
        
        # Check message text entities
        if hasattr(message, 'entities') and message.entities:
            for entity in message.entities:
                if entity.type in ["url", "text_link"]:
                    if hasattr(entity, 'url') and entity.url:
                        entity_urls.append(entity.url)
                    elif entity.type == "url":
                        # Extract URL from message text using entity offset and length
                        start = entity.offset
                        end = entity.offset + entity.length
                        entity_url = message_text[start:end]
                        if entity_url.startswith('http'):
                            entity_urls.append(entity_url)
        
        # Check caption entities (URLs might be embedded in captions!)
        if hasattr(message, 'caption_entities') and message.caption_entities:
            caption_text = message.caption or ""
            for entity in message.caption_entities:
                if entity.type in ["url", "text_link"]:
                    if hasattr(entity, 'url') and entity.url:
                        entity_urls.append(entity.url)
                        logger.info(f"Found URL in caption entity: {entity.url}")
                    elif entity.type == "url":
                        # Extract URL from caption using entity offset and length
                        start = entity.offset
                        end = entity.offset + entity.length
                        if start < len(caption_text) and end <= len(caption_text):
                            entity_url = caption_text[start:end]
                            if entity_url.startswith('http'):
                                entity_urls.append(entity_url)
                                logger.info(f"Extracted URL from caption entity: {entity_url}")
        
        # Log all entity information for debugging and CAPTURE URLs
        if hasattr(message, 'caption_entities') and message.caption_entities:
            logger.info(f"Caption entities found: {len(message.caption_entities)}")
            for i, entity in enumerate(message.caption_entities):
                logger.info(f"Entity {i}: type={entity.type}, offset={entity.offset}, length={entity.length}")
                if hasattr(entity, 'url') and entity.url:
                    logger.info(f"Entity {i} URL: {entity.url}")
                    # CRITICAL: Ensure all entity URLs are captured
                    if entity.url not in entity_urls:
                        entity_urls.append(entity.url)
                        logger.info(f"Added entity URL to list: {entity.url}")
                else:
                    logger.info(f"Entity {i} URL: None")
                    
        # CRITICAL: Check for URLs embedded in message reply_markup or buttons  
        if hasattr(message, 'reply_markup') and message.reply_markup:
            # Check if it's an InlineKeyboardMarkup with inline_keyboard
            if hasattr(message.reply_markup, 'inline_keyboard') and message.reply_markup.inline_keyboard:
                for row in message.reply_markup.inline_keyboard:
                    for button in row:
                        if hasattr(button, 'url') and button.url:
                            entity_urls.append(button.url)
                            logger.info(f"Found URL in reply markup button: {button.url}")
        
        # Combine all text sources for comprehensive extraction FIRST
        all_text_sources = [message_text] + entity_urls
        combined_text = " ".join(filter(None, all_text_sources))
        
        # ADVANCED: Try to decode potential hidden URLs in "info" sections
        # Some bots encode URLs within the "info" text using various methods
        if 'info' in combined_text.lower():
            logger.info(f"Processing 'info' section for hidden URLs...")
            
            # Method 1: Check if URL might be base64 encoded after "info"  
            import base64
            import re
            
            # Look for base64-like strings after "info"
            info_pattern = r'info[\s\xad\u2063]*([A-Za-z0-9+/=]{20,})'
            info_matches = re.findall(info_pattern, combined_text)
            
            for encoded_text in info_matches:
                try:
                    # Try to decode as base64
                    decoded_bytes = base64.b64decode(encoded_text + '==')  # Add padding
                    decoded_text = decoded_bytes.decode('utf-8')
                    if 'spotify' in decoded_text.lower():
                        logger.info(f"Found potential Spotify URL in base64: {decoded_text}")
                        entity_urls.append(decoded_text)
                except:
                    pass
            
            # Method 2: Check if there are invisible characters containing URLs
            # Remove only visible characters and see what's left
            visible_removed = ''.join(c for c in combined_text if ord(c) > 127)
            if len(visible_removed) > 10:
                logger.info(f"Found invisible character sequence: {repr(visible_removed)}")
                
            # Method 3: Look for URL patterns in character codes
            # Sometimes URLs are encoded as character sequences
            for char in combined_text:
                if ord(char) > 127:  # Non-ASCII character
                    logger.debug(f"Special character found: {repr(char)} (code: {ord(char)})")
            
            # Re-combine if new URLs were found
            if len(entity_urls) > len([message_text]):
                all_text_sources = [message_text] + entity_urls
                combined_text = " ".join(filter(None, all_text_sources))
        
        # Extract track info from combined sources
        track_info = extract_track_info(combined_text)
        
        # CRITICAL: If we found Spotify URLs in entities, use those directly  
        spotify_entity_urls = [url for url in entity_urls if 'spotify.com/track/' in url]
        if spotify_entity_urls:
            logger.info(f"Found {len(spotify_entity_urls)} Spotify URLs in entities: {spotify_entity_urls}")
            for url in spotify_entity_urls:
                direct_track_info = extract_track_info(url)
                logger.info(f"Direct extraction from {url}: {direct_track_info}")
                if direct_track_info and direct_track_info.get('track_id'):
                    track_info = direct_track_info
                    logger.info(f"SUCCESS: Using Spotify URL from entity: {url} -> Track ID: {direct_track_info.get('track_id')}")
                    break
        
        # Enhanced debug logging for track extraction
        if combined_text and ("spotify" in combined_text.lower() or "info" in combined_text.lower() or entity_urls):
            logger.info(f"Message text: {repr(message_text)}")
            logger.info(f"Entity URLs: {entity_urls}")
            logger.info(f"Combined text: {repr(combined_text)}")
            logger.info(f"Track extraction result: {track_info}")
            
            # If we have Spotify URLs but no track info, something is wrong
            spotify_urls = [url for url in entity_urls if 'spotify.com/track/' in url]
            if spotify_urls and not track_info.get('track_id'):
                logger.warning(f"Found Spotify URLs but no track extracted: {spotify_urls}")
                # Try to extract directly
                for url in spotify_urls:
                    logger.info(f"Attempting direct extraction from: {url}")
                    direct_result = extract_track_info(url)
                    logger.info(f"Direct extraction result: {direct_result}")
                    if direct_result and direct_result.get('track_id'):
                        track_info = direct_result
                        logger.info(f"SUCCESS: Using direct extraction result")
                        break
        
        # Forward file to backup channel with rate limiting
        backup_file_id = await forward_to_backup(client, message, track_info)
        
        # Get additional audio metadata if available
        audio_metadata = {}
        if message.audio:
            audio_metadata.update({
                "performer": message.audio.performer,
                "title": message.audio.title,
                "thumbnail": message.audio.thumbs[0].file_id if message.audio.thumbs else None
            })
        elif message.video:
            audio_metadata.update({
                "thumbnail": message.video.thumbs[0].file_id if message.video.thumbs else None
            })
        elif message.document and message.document.thumbs:
            audio_metadata.update({
                "thumbnail": message.document.thumbs[0].file_id
            })
        
        # Add rate limiting for media processing  
        await asyncio.sleep(0.1)  # Small delay to prevent rate limits
        
        # Prepare comprehensive document for MongoDB with TRACK INFORMATION
        document = {
            "file_id": file_data["file_id"],
            "backup_file_id": backup_file_id,
            "file_unique_id": file_data["file_unique_id"],
            "file_name": file_data.get("file_name"),
            "caption": message.caption or "",
            "file_type": file_data["file_type"],
            "mime_type": file_data.get("mime_type"),
            "file_size": file_data.get("file_size"),
            "duration": file_data.get("duration"),
            "width": file_data.get("width"),
            "height": file_data.get("height"),
            "chat_id": message.chat.id,
            "chat_title": message.chat.title or (message.chat.first_name if message.chat.first_name else "Unknown"),
            "message_id": message.id,
            "sender_id": message.from_user.id if message.from_user else None,
            "sender_username": message.from_user.username if message.from_user else None,
            "sender_first_name": message.from_user.first_name if message.from_user else None,
            "sender_last_name": message.from_user.last_name if message.from_user else None,
            "date": message.date.isoformat(),
            "is_deleted": False,
            
            # CRITICAL: Include track information in database
            "track_url": track_info.get("track_url") if track_info else None,
            "track_id": track_info.get("track_id") if track_info else None,
            "platform": track_info.get("platform") if track_info else None,
            "title": track_info.get("title") if track_info else None,
            "artist": track_info.get("artist") if track_info else None,
            **audio_metadata  # Include performer, title, thumbnail if available
        }
        
        # Store in MongoDB
        result = db.insert_file(document)
        
        if result:
            logger.info(f"Successfully indexed file: {file_data['file_id']}")
        else:
            logger.error(f"Failed to index file: {file_data['file_id']}")
            
    except Exception as e:
        logger.error(f"Error handling media message: {e}")

async def forward_to_backup(client: Client, message: Message, track_info: dict = None):
    """Forward message to backup channel with proper caption and return backup file_id"""
    try:
        backup_channel_id = db.get_backup_channel_id()
        if not backup_channel_id:
            logger.warning("No backup channel configured")
            return None
            
        # Use provided track_info or extract from message
        if not track_info:
            message_text = message.text or message.caption or ""
            track_info = extract_track_info(message_text)
        
        # Check if this file already exists in backup channel to prevent duplicates
        file_data = get_file_metadata(message)
        if file_data:
            try:
                existing_backup = db.get_file_by_backup_id(file_data["file_id"])
                if existing_backup:
                    logger.info(f"File already in backup channel, skipping forward: {file_data['file_id']}")
                    return existing_backup.get("backup_file_id")
            except AttributeError:
                logger.debug("Backup check method not available, proceeding with forward")
        
        # Create minimal caption with title, artist, and track ID
        backup_caption = format_file_caption(message, minimal_format=True, track_info_override=track_info)
        
        # Send file to backup channel with proper caption based on file type
        forwarded_msg = None
        
        try:
            if message.audio:
                forwarded_msg = await client.send_audio(
                    chat_id=backup_channel_id,
                    audio=message.audio.file_id,
                    caption=backup_caption,
                    duration=message.audio.duration,
                    performer=message.audio.performer,
                    title=message.audio.title
                )
            elif message.video:
                forwarded_msg = await client.send_video(
                    chat_id=backup_channel_id,
                    video=message.video.file_id,
                    caption=backup_caption,
                    duration=message.video.duration,
                    width=message.video.width,
                    height=message.video.height
                )
            elif message.document:
                forwarded_msg = await client.send_document(
                    chat_id=backup_channel_id,
                    document=message.document.file_id,
                    caption=backup_caption,
                    file_name=message.document.file_name
                )
            elif message.photo:
                forwarded_msg = await client.send_photo(
                    chat_id=backup_channel_id,
                    photo=message.photo.file_id,
                    caption=backup_caption
                )
            
            # Add rate limiting delay to avoid FloodWait errors
            await asyncio.sleep(0.5)  # Small delay between operations
                
        except FloodWait as e:
            logger.warning(f"Rate limit hit, waiting {e.value} seconds")
            await asyncio.sleep(float(e.value))
            # Retry the operation
            return await forward_to_backup(client, message)
        
        if forwarded_msg:
            backup_file_data = get_file_metadata(forwarded_msg)
            return backup_file_data["file_id"] if backup_file_data else None
            
    except Exception as e:
        logger.error(f"Error sending to backup channel: {e}")
        return None

async def handle_send_command(client: Client, message: Message):
    """Handle /send command to retrieve file by filename"""
    try:
        # Extract filename from command
        command_parts = message.text.split(" ", 1)
        if len(command_parts) < 2:
            await message.reply("Usage: /send <filename>")
            return
            
        filename = command_parts[1].strip()
        
        # Search for file in database
        file_doc = db.find_file_by_name(filename)
        
        if not file_doc:
            await message.reply(f"File '{filename}' not found in database.")
            return
            
        # Send the file from backup
        await send_file_from_backup(client, message, file_doc)
        
    except Exception as e:
        logger.error(f"Error handling send command: {e}")
        await message.reply("An error occurred while processing your request.")

async def handle_sendid_command(client: Client, message: Message):
    """Handle /sendid command to retrieve file by track ID"""
    try:
        # Extract track ID from command
        command_parts = message.text.split(" ", 1)
        if len(command_parts) < 2:
            await message.reply("Usage: /sendid <track_id>")
            return
            
        track_id = command_parts[1].strip()
        
        # Search for file in database
        file_doc = db.find_file_by_track_id(track_id)
        
        if not file_doc:
            await message.reply(f"Track ID '{track_id}' not found in database.")
            return
            
        # Send the file from backup
        await send_file_from_backup(client, message, file_doc)
        
    except Exception as e:
        logger.error(f"Error handling sendid command: {e}")
        await message.reply("An error occurred while processing your request.")

async def send_file_from_backup(client: Client, message: Message, file_doc: dict):
    """Send file from backup with formatted caption"""
    try:
        backup_file_id = file_doc.get("backup_file_id")
        if not backup_file_id:
            await message.reply("Backup file not available for this item.")
            return
            
        # Generate formatted caption
        caption = format_file_caption(file_doc)
        
        # Determine file type and send accordingly
        file_type = file_doc.get("file_type")
        
        if file_type == "audio":
            await client.send_audio(
                chat_id=message.chat.id,
                audio=backup_file_id,
                caption=caption
            )
        elif file_type == "video":
            await client.send_video(
                chat_id=message.chat.id,
                video=backup_file_id,
                caption=caption
            )
        elif file_type == "document":
            await client.send_document(
                chat_id=message.chat.id,
                document=backup_file_id,
                caption=caption
            )
        elif file_type == "photo":
            await client.send_photo(
                chat_id=message.chat.id,
                photo=backup_file_id,
                caption=caption
            )
        else:
            await message.reply("Unsupported file type for retrieval.")
            
    except Exception as e:
        logger.error(f"Error sending file from backup: {e}")
        await message.reply("Failed to retrieve and send the file.")

async def handle_start_command(client: Client, message: Message):
    """Handle /start command"""
    global indexing_process
    
    welcome_text = """
🤖 **Media Indexer Bot**

This bot automatically indexes media files and provides retrieval functionality with complete metadata extraction.

**Commands:**
• `/send <filename>` - Retrieve file by filename
• `/sendid <track_id>` - Retrieve file by track ID
• `/stats` - Show database statistics
• `/db` - Export database as PDF (key fields)
• `/db excel` - Export database as Excel with ALL 30+ metadata fields
• `/db csv` - Export database as CSV with ALL 30+ metadata fields
• `/cancel` - Stop current indexing process

**To start indexing:**
1. For private channels: Add this bot as admin to the channel
2. For public channels: No admin access needed
3. Send any message link (t.me/channel/123) or forward any message from the channel
4. Bot will automatically start indexing from that message and show progress

**Features:**
✅ Extracts Spotify track IDs from TEXT_LINK entities
✅ Processes invisible separator characters (\\u2063, \\xad)
✅ Stores 30+ metadata fields per file
✅ Automatic backup to designated channel
✅ Real-time progress tracking with rate limiting
    """
    await message.reply(welcome_text)
    
    # Auto-start indexing from the specified channel if not already running
    if not indexing_process["active"]:
        logger.info("Auto-starting indexing from Spotifyapk56 channel")
        await auto_start_spotifyapk_indexing(client, message)

async def handle_stats_command(client: Client, message: Message):
    """Handle /stats command to show database statistics"""
    try:
        stats = db.get_statistics()
        
        stats_text = f"""
📊 **Database Statistics**

• Total Files: {stats['total_files']}
• Audio Files: {stats['audio_files']}
• Video Files: {stats['video_files']}
• Documents: {stats['document_files']}
• Photos: {stats['photo_files']}
• Files with Track URLs: {stats['files_with_tracks']}
        """
        
        await message.reply(stats_text)
        
    except Exception as e:
        logger.error(f"Error getting statistics: {e}")
        await message.reply("Failed to retrieve statistics.")

async def handle_message_link(client: Client, message: Message):
    """Handle message links to start indexing"""
    global indexing_process
    
    if indexing_process["active"]:
        await message.reply("⚠️ Indexing is already in progress.")
        return
    
    try:
        text = message.text or message.caption or ""
        
        # Extract message link
        link_pattern = r"https://t\.me/([^/]+)/(\d+)"
        match = re.search(link_pattern, text)
        
        if match:
            channel_username = match.group(1)
            message_id = int(match.group(2))
            
            # Try to get chat info
            try:
                chat = await client.get_chat(channel_username)
                # Handle different chat types
                if hasattr(chat, 'id'):
                    chat_id = chat.id
                else:
                    # For ChatPreview objects, try to extract ID from username
                    try:
                        chat_full = await client.get_chat(f"@{channel_username}")
                        chat_id = chat_full.id
                    except:
                        await message.reply(f"❌ Cannot access channel @{channel_username}. Please make sure the bot has access.")
                        return
                
                chat_title = getattr(chat, 'title', None) or getattr(chat, 'username', None) or channel_username
                
                # Check if there's previous indexing progress
                last_indexed = db.get_stored_last_indexed_message_id(chat_id)
                
                if last_indexed > 1:
                    resume_start = last_indexed + 1
                    await message.reply(f"🔄 Found previous progress for {chat_title}\n"
                                      f"Last indexed: Message {last_indexed}\n"
                                      f"Will resume from: Message {resume_start}\n"
                                      f"Target: Message {message_id}")
                    await start_indexing_process(client, message, chat_id, resume_start, chat_title, message_id)
                else:
                    await start_indexing_process(client, message, chat_id, 1, chat_title, message_id)
                
            except ChannelPrivate:
                await message.reply(f"""
❌ **Channel is Private**

To index this channel:
1. Add this bot as **admin** to @{channel_username}
2. Give it permission to read messages
3. Send the message link again
                """)
                
            except ChatAdminRequired:
                await message.reply(f"""
❌ **Admin Rights Required**

To index this channel:
1. Add this bot as **admin** to @{channel_username}
2. Give it permission to read messages
3. Send the message link again
                """)
                
            except UsernameNotOccupied:
                await message.reply(f"❌ Channel @{channel_username} not found.")
                
    except Exception as e:
        logger.error(f"Error handling message link: {e}")
        await message.reply("❌ Error processing message link.")



async def start_indexing_process(client: Client, message: Message, chat_id: int, start_message_id: int, chat_title: str = "Unknown", target_message_id: int = None):
    """Start the indexing process for a channel"""
    global indexing_process
    
    try:
        # Check if bot has access to the channel
        try:
            chat = await client.get_chat(chat_id)
            if start_message_id is None:
                # Check for last indexed message to resume from
                last_indexed = db.get_stored_last_indexed_message_id(chat_id)
                start_message_id = last_indexed + 1 if last_indexed > 1 else 1
                    
        except Exception as e:
            if "CHAT_ADMIN_REQUIRED" in str(e):
                await message.reply(f"""
❌ **Bot needs admin access**

To index **{chat_title}**:
1. Add this bot as admin to the channel
2. Give it permission to read messages
3. Try again
                """)
                return
            else:
                raise e
        
        # Use target_message_id or start_message_id as end point
        end_message_id = target_message_id or start_message_id
        
        # Initialize indexing process
        indexing_process.update({
            "active": True,
            "chat_id": chat_id,
            "message_id": start_message_id,
            "total": 0,
            "processed": 0,
            "stop_requested": False
        })
        
        # Send initial status
        last_indexed = db.get_stored_last_indexed_message_id(chat_id)
        resume_text = f"\n🔄 **Resuming from:** Message {last_indexed + 1} (last indexed: {last_indexed})" if last_indexed > 1 else ""
        
        status_msg = await message.reply(f"""
🚀 **Starting Indexing Process**

📂 **Channel:** {chat_title}
🔍 **Range:** Message {start_message_id} to {end_message_id}{resume_text}

⏳ Scanning messages...
        """)
        
        # Start indexing in background
        asyncio.create_task(index_channel_messages(client, status_msg, chat_id, start_message_id, chat_title, end_message_id))
        
    except Exception as e:
        logger.error(f"Error starting indexing: {e}")
        await message.reply("❌ Failed to start indexing process.")

async def index_channel_messages(client: Client, status_msg: Message, chat_id: int, start_message_id: int, chat_title: str, end_message_id: int = None):
    """Index messages from a channel with progress updates and progress saving"""
    global indexing_process
    
    try:
        # Use end_message_id if provided, otherwise use start_message_id
        final_message_id = end_message_id or start_message_id
        
        # Set initial estimate
        indexing_process["total"] = final_message_id - start_message_id + 1
        
        await status_msg.edit_text(f"""
🚀 **Indexing Process Started**

📂 **Channel:** {chat_title}
🔍 **Range:** Message {start_message_id} to {final_message_id}

⏳ Searching for media files...
        """)
        
        # Initialize tracking variables
        total_messages = final_message_id - start_message_id + 1
        fetched_messages = 0
        processed = 0
        errors = 0
        current_msg_id = start_message_id  # Start from provided start message ID
        consecutive_failures = 0
        max_failures = 50  # Allow more failures before stopping
        # Removed max_processed limit to allow unlimited processing
        last_update_time = time.time()  # Track last progress update time
        last_saved_progress = start_message_id  # Track last saved progress
        start_time = time.time()  # Track start time for speed calculation
        
        # Rate limiting: 20 messages per minute = 1 message every 3 seconds
        rate_limit_delay = 3.0  # seconds between messages
        last_message_time = 0
        
        await status_msg.edit_text(f"""
🚀 **Indexing Process Started**

📂 **Channel:** {chat_title}
🔍 **Range:** Message {start_message_id} to {final_message_id}
📊 **Total messages to check:** {total_messages}

⏳ Searching for media files...
        """)
        
        while consecutive_failures < max_failures and current_msg_id <= final_message_id and not indexing_process["stop_requested"]:
            try:
                # Try to get the specific message
                try:
                    messages = await client.get_messages(chat_id, current_msg_id)
                    consecutive_failures = 0  # Reset failure count on success
                    fetched_messages += 1
                    
                    # Handle both single message and list of messages
                    msg_list = messages if isinstance(messages, list) else [messages]
                    
                    for msg in msg_list:
                        if msg and (msg.audio or msg.video or msg.document or msg.photo):
                            try:
                                # Apply rate limiting: 20 messages per minute (3 seconds between messages)
                                current_time = time.time()
                                time_since_last = current_time - last_message_time
                                if time_since_last < rate_limit_delay:
                                    sleep_time = rate_limit_delay - time_since_last
                                    logger.info(f"Rate limiting: sleeping for {sleep_time:.1f} seconds")
                                    await asyncio.sleep(sleep_time)
                                
                                await handle_media_message(client, msg)
                                processed += 1
                                indexing_process["processed"] = processed
                                last_message_time = time.time()
                                
                                # Update progress every 2 minutes or every 20 files (whichever comes first)
                                current_time = time.time()
                                time_since_update = current_time - last_update_time
                                if time_since_update >= 120 or processed % 20 == 0:  # 120 seconds = 2 minutes
                                    last_update_time = current_time
                                    # Calculate progress percentage based on current position
                                    progress_percentage = min(100, int(((current_msg_id - start_message_id + 1) / total_messages) * 100))
                                    
                                    # Create fancy status with proper progress bar and speed
                                    fancy_status = create_fancy_progress_status(
                                        processed=processed,
                                        errors=errors,
                                        current_msg_id=current_msg_id,
                                        chat_title=chat_title,
                                        total_messages=total_messages,
                                        fetched_messages=fetched_messages,
                                        skipped=fetched_messages - processed,
                                        percentage=progress_percentage,
                                        start_time=start_time,
                                        start_msg_id=start_message_id
                                    )
                                    
                                    await status_msg.edit_text(f"```\n{fancy_status}\n```")
                                    
                            except Exception as e:
                                logger.error(f"Error processing message {msg.id}: {e}")
                                errors += 1
                    
                    # Save progress every 50 messages to allow resuming
                    if current_msg_id - last_saved_progress >= 50:
                        db.update_last_indexed_message_id(chat_id, current_msg_id)
                        last_saved_progress = current_msg_id
                        logger.info(f"Progress saved: last indexed message {current_msg_id}")
                            
                except Exception as e:
                    # Message doesn't exist or can't be accessed
                    consecutive_failures += 1
                    if "MESSAGE_ID_INVALID" not in str(e):
                        logger.debug(f"Could not get message {current_msg_id}: {e}")
                
                # Move to next message
                current_msg_id += 1
                    
            except Exception as e:
                logger.error(f"Error in message iteration loop: {e}")
                consecutive_failures += 1
                current_msg_id += 1
                
        # Save final progress
        final_processed_id = current_msg_id - 1
        db.update_last_indexed_message_id(chat_id, final_processed_id)
        logger.info(f"Final progress saved: last indexed message {final_processed_id}")
        
        # Final status
        if indexing_process["stop_requested"]:
            await status_msg.edit_text(f"""
⚠️ **Indexing Stopped**

📂 **Channel:** {chat_title}
📊 **Processed:** {processed} media files
📍 **Last indexed:** Message {final_processed_id}
❌ **Stopped by user**

✅ Progress saved - will resume from message {final_processed_id + 1} on next start
            """)
        else:
            total_fetched = current_msg_id - start_message_id
            skipped_final = total_fetched - processed
            final_status = create_final_status(
                processed=processed,
                errors=errors,
                chat_title=chat_title,
                total_messages=total_messages,
                fetched_messages=total_fetched,
                skipped=skipped_final
            )
            
            # Add progress saved info to final status
            final_status += f"\n\n✅ Progress saved - last indexed: {final_processed_id}"
            
            await status_msg.edit_text(f"```\n{final_status}\n```")
            
    except Exception as e:
        logger.error(f"Error during indexing: {e}")
        processed = indexing_process.get("processed", 0)
        await status_msg.edit_text(f"""
❌ **Indexing Failed**

📂 **Channel:** {chat_title}
📊 **Processed:** {processed} media files
❌ **Error:** {str(e)}
        """)
    finally:
        indexing_process["active"] = False

def create_fancy_progress_status(processed: int, errors: int, current_msg_id: int, chat_title: str, total_messages: int, fetched_messages: int, skipped: int = 0, percentage: int = 0, start_time: float = None, start_msg_id: int = None) -> str:
    """Create a fancy progress status display with speed calculation"""
    
    # Calculate correct percentage based on message progress
    if start_msg_id:
        progress_made = current_msg_id - start_msg_id + 1
        percentage = int((progress_made / total_messages) * 100) if total_messages > 0 else 0
    elif percentage == 0:
        percentage = int((fetched_messages / total_messages) * 100) if total_messages > 0 else 0
    
    # Calculate speed
    speed_text = "Calculating..."
    if start_time and processed > 0:
        elapsed_time = time.time() - start_time
        if elapsed_time > 0:
            files_per_sec = processed / elapsed_time
            files_per_min = files_per_sec * 60
            speed_text = f"{files_per_min:.1f} files/min"
    
    status_text = f"""╔════❰ ɪɴᴅᴇxɪɴɢ sᴛᴀᴛᴜs  ❱═❍⊱❁
║╭━━━━━━━━━━━━━━━➣
║┣⪼𖨠 ᴄʜᴀɴɴᴇʟ ɴᴀᴍᴇ:  {chat_title}
║┃
║┣⪼𖨠 ᴛᴏᴛᴀʟ ᴍᴇssᴀɢᴇs:  {total_messages}
║┃
║┣⪼𖨠 ғᴇᴛᴄʜᴇᴅ ᴍᴇssᴀɢᴇs:  {fetched_messages}
║┃
║┣⪼𖨠 ɪɴᴅᴇxᴇᴅ ᴍᴇᴅɪᴀ:  {processed}
║┃
║┣⪼𖨠 ᴇʀʀᴏʀ ᴄᴏᴜɴᴛ:  {errors}
║┃
║┣⪼𖨠 sᴋɪᴘᴘᴇᴅ ᴍᴇssᴀɢᴇs:  {skipped}
║┃
║┣⪼𖨠 ᴄᴜʀʀᴇɴᴛ ᴍᴇssᴀɢᴇ:  {current_msg_id}
║┃
║┣⪼𖨠 ᴘʀᴏᴄᴇssɪɴɢ sᴘᴇᴇᴅ:  {speed_text}
║┃
║┣⪼𖨠 ᴘᴇʀᴄᴇɴᴛᴀɢᴇ:  {percentage}%
║╰━━━━━━━━━━━━━━━➣ 
╚════❰ ᴘʀᴏᴄᴇssɪɴɢ ❱══❍⊱❁"""
    
    return status_text

def create_final_status(processed: int, errors: int, chat_title: str, total_messages: int, fetched_messages: int, skipped: int = 0) -> str:
    """Create final completion status"""
    
    status_text = f"""╔════❰ ɪɴᴅᴇxɪɴɢ ᴄᴏᴍᴘʟᴇᴛᴇ  ❱═❍⊱❁
║╭━━━━━━━━━━━━━━━➣
║┣⪼𖨠 ᴄʜᴀɴɴᴇʟ ɴᴀᴍᴇ:  {chat_title}
║┃
║┣⪼𖨠 ᴛᴏᴛᴀʟ ᴍᴇssᴀɢᴇs:  {total_messages}
║┃
║┣⪼𖨠 ғᴇᴛᴄʜᴇᴅ ᴍᴇssᴀɢᴇs:  {fetched_messages}
║┃
║┣⪼𖨠 ɪɴᴅᴇxᴇᴅ ᴍᴇᴅɪᴀ:  {processed}
║┃
║┣⪼𖨠 ᴇʀʀᴏʀ ᴄᴏᴜɴᴛ:  {errors}
║┃
║┣⪼𖨠 sᴋɪᴘᴘᴇᴅ ᴍᴇssᴀɢᴇs:  {skipped}
║┃
║┣⪼𖨠 ᴄᴜʀʀᴇɴᴛ sᴛᴀᴛᴜs:  ᴄᴏᴍᴘʟᴇᴛᴇᴅ
║┃
║┣⪼𖨠 ᴘᴇʀᴄᴇɴᴛᴀɢᴇ:  100%
║╰━━━━━━━━━━━━━━━➣ 
╚════❰ ғɪɴɪsʜᴇᴅ ❱══❍⊱❁

Use /send <filename> or /sendid <track_id> to retrieve files."""
    
    return status_text

async def handle_stop_index_command(client: Client, message: Message):
    """Handle /stop_index command"""
    global indexing_process
    
    if not indexing_process["active"]:
        await message.reply("ℹ️ No indexing process is currently running.")
        return
        
    indexing_process["stop_requested"] = True
    await message.reply("⚠️ Stopping indexing process...")

async def handle_db_command(client: Client, message: Message):
    """Handle /db command to export database as PDF and Excel with ALL metadata fields"""
    try:
        command_parts = message.text.split()
        export_format = "pdf"  # Default
        
        if len(command_parts) > 1:
            export_format = command_parts[1].lower()
            if export_format not in ["pdf", "excel", "xlsx", "csv"]:
                await message.reply("📊 Usage: /db [pdf|excel|xlsx|csv]\nDefault: pdf")
                return
        
        if export_format in ["excel", "xlsx"]:
            format_name = "Excel"
        elif export_format == "csv":
            format_name = "CSV"
        else:
            format_name = "PDF"
        await message.reply(f"📊 Generating database export as {format_name} with ALL metadata fields... This may take a few moments.")
        
        # Get all files from database
        files = db.get_all_files()
        
        if not files:
            await message.reply("📁 Database is empty. No files to export.")
            return
        
        # Import libraries based on format
        if export_format in ["excel", "xlsx"]:
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                from datetime import datetime
                import tempfile
                import os
            except ImportError as e:
                await message.reply(f"❌ Excel generation libraries not available: {e}")
                return
        else:
            try:
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch
                from datetime import datetime
                import tempfile
                import os
            except ImportError as e:
                await message.reply(f"❌ PDF generation libraries not available: {e}")
                return
        
        if export_format in ["excel", "xlsx"]:
            # Generate Excel with ALL metadata fields
            await generate_excel_export(client, message, files)
        elif export_format == "csv":
            # Generate CSV with ALL metadata fields
            await generate_csv_export(client, message, files)
        else:
            # Generate PDF with ALL metadata fields  
            await generate_pdf_export(client, message, files)
            
    except Exception as e:
        logger.error(f"Error handling db command: {e}")
        await message.reply(f"❌ Failed to export database: {str(e)}")

async def generate_excel_export(client: Client, message: Message, files):
    """Generate comprehensive Excel export with ALL 30+ metadata fields"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import tempfile
        import os
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            excel_path = tmp_file.name
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Media Index Database"
        
        # ALL metadata fields (30+ fields)
        headers = [
            '#', 'File ID', 'Backup File ID', 'File Unique ID', 'File Name', 'Caption', 
            'File Type', 'MIME Type', 'File Size (MB)', 'Duration (sec)', 'Width', 'Height',
            'Chat ID', 'Chat Title', 'Message ID', 'Sender ID', 'Sender Username', 
            'Sender First Name', 'Sender Last Name', 'Date', 'Is Deleted',
            'Track URL', 'Track ID', 'Platform', 'Performer', 'Title', 'Thumbnail',
            'Original Caption', 'Entity URLs', 'Processing Status'
        ]
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_idx, file_doc in enumerate(files, 2):
            if not file_doc:
                continue
                
            # Extract all fields with safe defaults
            file_size = file_doc.get('file_size', 0) or 0
            size_mb = round(file_size / (1024 * 1024), 2) if file_size else 0
            
            row_data = [
                row_idx - 1,  # Row number
                file_doc.get('file_id', ''),
                file_doc.get('backup_file_id', ''),
                file_doc.get('file_unique_id', ''),
                file_doc.get('file_name', ''),
                file_doc.get('caption', ''),
                file_doc.get('file_type', ''),
                file_doc.get('mime_type', ''),
                size_mb,
                file_doc.get('duration', 0) or 0,
                file_doc.get('width', 0) or 0,
                file_doc.get('height', 0) or 0,
                file_doc.get('chat_id', ''),
                file_doc.get('chat_title', ''),
                file_doc.get('message_id', ''),
                file_doc.get('sender_id', ''),
                file_doc.get('sender_username', ''),
                file_doc.get('sender_first_name', ''),
                file_doc.get('sender_last_name', ''),
                file_doc.get('date', ''),
                file_doc.get('is_deleted', False),
                file_doc.get('track_url', ''),
                file_doc.get('track_id', ''),
                file_doc.get('platform', ''),
                file_doc.get('performer', ''),
                file_doc.get('title', ''),
                file_doc.get('thumbnail', ''),
                file_doc.get('original_caption', ''),
                str(file_doc.get('entity_urls', [])),
                'Processed' if file_doc.get('track_id') else 'No Track'
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add statistics sheet
        stats_ws = wb.create_sheet("Statistics")
        stats = db.get_statistics()
        
        stats_data = [
            ['Database Statistics', ''],
            ['Total Files', stats['total_files']],
            ['Audio Files', stats['audio_files']],
            ['Video Files', stats['video_files']],
            ['Document Files', stats['document_files']], 
            ['Photo Files', stats['photo_files']],
            ['Files with Track URLs', stats['files_with_tracks']],
            ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['Field Descriptions', ''],
            ['Track URL', 'Full Spotify/Music URL'],
            ['Track ID', 'Extracted music track identifier'],
            ['Platform', 'Music platform (spotify, youtube, etc.)'],
            ['File Size (MB)', 'File size in megabytes'],
            ['Duration (sec)', 'Audio/video duration in seconds'],
            ['Chat Title', 'Source channel/chat name'],
            ['Processing Status', 'Whether track info was extracted']
        ]
        
        for row_idx, (key, value) in enumerate(stats_data, 1):
            stats_ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            stats_ws.cell(row=row_idx, column=2, value=value)
        
        # Save workbook
        wb.save(excel_path)
        
        # Send Excel file
        await client.send_document(
            chat_id=message.chat.id,
            document=excel_path,
            caption=f"📊 **Complete Database Export (Excel)**\n\n"
                   f"📁 **{stats['total_files']} files** with **ALL {len(headers)} metadata fields**\n"
                   f"🎵 **{stats['files_with_tracks']} files** have track information\n"
                   f"📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                   f"**Includes:** All file metadata, track IDs, Spotify URLs, sender info, timestamps, and more!"
        )
        
        # Clean up
        os.unlink(excel_path)
        await message.reply("✅ Excel export completed with ALL metadata fields!")
        
    except Exception as e:
        logger.error(f"Error generating Excel export: {e}")
        await message.reply(f"❌ Failed to generate Excel export: {str(e)}")

async def generate_pdf_export(client: Client, message: Message, files):
    """Generate comprehensive PDF export with ALL metadata fields"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from datetime import datetime
        import tempfile
        import os
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            pdf_path = tmp_file.name
        
        # Use landscape orientation for more columns
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4), 
                               rightMargin=36, leftMargin=36, 
                               topMargin=36, bottomMargin=36)
        
        # Container for PDF elements
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        title = Paragraph("📊 Complete Media Database Export - ALL FIELDS", title_style)
        story.append(title)
        
        # Statistics
        stats = db.get_statistics()
        stats_text = f"""
        <b>Database Statistics:</b><br/>
        • Total Files: {stats['total_files']}<br/>
        • Audio Files: {stats['audio_files']}<br/>
        • Video Files: {stats['video_files']}<br/>
        • Document Files: {stats['document_files']}<br/>
        • Photo Files: {stats['photo_files']}<br/>
        • Files with Track URLs: {stats['files_with_tracks']}<br/>
        • Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        stats_para = Paragraph(stats_text, styles['Normal'])
        story.append(stats_para)
        story.append(Spacer(1, 20))
        
        # Prepare table data with key fields (PDF has space constraints)
        data = [['#', 'File Name', 'Type', 'Size (MB)', 'Track ID', 'Track URL', 'Platform', 'Chat Title', 'Date']]
        
        for i, file_doc in enumerate(files[:500], 1):  # Limit for PDF readability
            if not file_doc:
                continue
                
            file_name = str(file_doc.get('file_name', 'Unknown'))[:25]  # Truncate for PDF
            file_type = str(file_doc.get('file_type', 'Unknown'))
            file_size = file_doc.get('file_size', 0) or 0
            size_mb = f"{round(file_size / (1024 * 1024), 2):.1f}" if file_size else "0"
            track_id = str(file_doc.get('track_id', ''))[:15]  # Truncate for PDF
            track_url = str(file_doc.get('track_url', ''))[:30]  # Truncate for PDF
            platform = str(file_doc.get('platform', ''))
            chat_title = str(file_doc.get('chat_title', ''))[:20]  # Truncate for PDF
            date = str(file_doc.get('date', ''))[:10]  # Date only
            
            data.append([i, file_name, file_type, size_mb, track_id, track_url, platform, chat_title, date])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Add note about complete data
        note_text = f"""
        <b>Note:</b> This PDF contains key fields from {len(files)} files. 
        For COMPLETE data with all 30+ metadata fields, use: <b>/db excel</b>
        """
        note_para = Paragraph(note_text, styles['Normal'])
        story.append(Spacer(1, 20))
        story.append(note_para)
        
        # Build PDF
        doc.build(story)
        
        # Send PDF file
        await client.send_document(
            chat_id=message.chat.id,
            document=pdf_path,
            caption=f"📊 **Database Export (PDF)**\n\n"
                   f"📁 **{stats['total_files']} files** with key metadata fields\n"
                   f"🎵 **{stats['files_with_tracks']} files** have track information\n"
                   f"📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                   f"💡 **For ALL 30+ fields, use:** `/db excel`"
        )
        
        # Clean up
        os.unlink(pdf_path)
        await message.reply("✅ PDF export completed! Use `/db excel` for complete data.")
        
    except Exception as e:
        logger.error(f"Error generating PDF export: {e}")
        await message.reply(f"❌ Failed to generate PDF export: {str(e)}")

async def generate_csv_export(client: Client, message: Message, files):
    """Generate comprehensive CSV export with ALL 30+ metadata fields"""
    try:
        import csv
        from datetime import datetime
        import tempfile
        import os
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='w', newline='', encoding='utf-8') as tmp_file:
            csv_path = tmp_file.name
            writer = csv.writer(tmp_file)
            
            # ALL metadata fields (30+ fields)
            headers = [
                'Row_Number', 'File_ID', 'Backup_File_ID', 'File_Unique_ID', 'File_Name', 'Caption', 
                'File_Type', 'MIME_Type', 'File_Size_MB', 'Duration_Sec', 'Width', 'Height',
                'Chat_ID', 'Chat_Title', 'Message_ID', 'Sender_ID', 'Sender_Username', 
                'Sender_First_Name', 'Sender_Last_Name', 'Date', 'Is_Deleted',
                'Track_URL', 'Track_ID', 'Platform', 'Performer', 'Title', 'Thumbnail',
                'Original_Caption', 'Entity_URLs', 'Processing_Status'
            ]
            
            # Write headers
            writer.writerow(headers)
            
            # Write data rows
            for row_idx, file_doc in enumerate(files, 1):
                if not file_doc:
                    continue
                    
                # Extract all fields with safe defaults
                file_size = file_doc.get('file_size', 0) or 0
                size_mb = round(file_size / (1024 * 1024), 2) if file_size else 0
                
                row_data = [
                    row_idx,  # Row number
                    file_doc.get('file_id', ''),
                    file_doc.get('backup_file_id', ''),
                    file_doc.get('file_unique_id', ''),
                    file_doc.get('file_name', ''),
                    file_doc.get('caption', ''),
                    file_doc.get('file_type', ''),
                    file_doc.get('mime_type', ''),
                    size_mb,
                    file_doc.get('duration', 0) or 0,
                    file_doc.get('width', 0) or 0,
                    file_doc.get('height', 0) or 0,
                    file_doc.get('chat_id', ''),
                    file_doc.get('chat_title', ''),
                    file_doc.get('message_id', ''),
                    file_doc.get('sender_id', ''),
                    file_doc.get('sender_username', ''),
                    file_doc.get('sender_first_name', ''),
                    file_doc.get('sender_last_name', ''),
                    file_doc.get('date', ''),
                    file_doc.get('is_deleted', False),
                    file_doc.get('track_url', ''),
                    file_doc.get('track_id', ''),
                    file_doc.get('platform', ''),
                    file_doc.get('performer', ''),
                    file_doc.get('title', ''),
                    file_doc.get('thumbnail', ''),
                    file_doc.get('original_caption', ''),
                    str(file_doc.get('entity_urls', [])),
                    'Processed' if file_doc.get('track_id') else 'No Track'
                ]
                
                writer.writerow(row_data)
        
        # Get statistics
        stats = db.get_statistics()
        
        # Send CSV file
        await client.send_document(
            chat_id=message.chat.id,
            document=csv_path,
            caption=f"📊 **Complete Database Export (CSV)**\n\n"
                   f"📁 **{stats['total_files']} files** with **ALL {len(headers)} metadata fields**\n"
                   f"🎵 **{stats['files_with_tracks']} files** have track information\n"
                   f"📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                   f"**Perfect for:** Data analysis, spreadsheet import, automated processing\n"
                   f"**To retrieve this file:** Use `/sendid` with the track ID or `/send` with filename"
        )
        
        # Store CSV file in database for retrieval
        try:
            # Read the file and store it in database with a unique identifier
            csv_filename = f"database_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            # Send to backup channel and get file ID
            backup_msg = await client.send_document(
                chat_id=os.getenv('BACKUP_CHANNEL_ID'),
                document=csv_path,
                caption=f"📊 Database Export CSV - {csv_filename}\n"
                       f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                       f"Files: {stats['total_files']} | Fields: {len(headers)}"
            )
            
            # Store in database for easy retrieval
            export_doc = {
                'file_id': backup_msg.document.file_id,
                'backup_file_id': backup_msg.document.file_id,
                'file_unique_id': backup_msg.document.file_unique_id,
                'file_name': csv_filename,
                'file_type': 'document',
                'mime_type': 'text/csv',
                'file_size': backup_msg.document.file_size,
                'chat_id': message.chat.id,
                'chat_title': 'Database Export',
                'message_id': backup_msg.id,
                'sender_id': message.from_user.id if message.from_user else None,
                'date': datetime.now().isoformat(),
                'caption': f"Database export with {stats['total_files']} files",
                'track_id': f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                'platform': 'database_export',
                'is_deleted': False
            }
            
            db.insert_file(export_doc)
            
            await message.reply(f"✅ CSV export completed!\n\n"
                              f"**To retrieve this file later:**\n"
                              f"• Use: `/send {csv_filename}`\n"
                              f"• Or: `/sendid export_{datetime.now().strftime('%Y%m%d_%H%M%S')}`")
            
        except Exception as store_error:
            logger.warning(f"Could not store CSV in database: {store_error}")
            await message.reply("✅ CSV export completed with ALL metadata fields!")
        
        # Clean up
        os.unlink(csv_path)
        
    except Exception as e:
        logger.error(f"Error generating CSV export: {e}")
        await message.reply(f"❌ Failed to generate CSV export: {str(e)}")

async def handle_forwarded_message(client: Client, message: Message):
    """Handle forwarded messages to automatically start indexing"""
    global indexing_process
    
    if message.forward_from_chat and not indexing_process["active"]:
        chat_id = message.forward_from_chat.id
        start_message_id = message.forward_from_message_id
        chat_title = message.forward_from_chat.title or message.forward_from_chat.username
        
        await start_indexing_process(client, message, chat_id, start_message_id, chat_title)

async def auto_start_spotifyapk_indexing(client: Client, message: Message):
    """Auto-start indexing from Spotifyapk56 channel starting from message 18149251"""
    global indexing_process
    
    try:
        # The specific channel
        channel_username = "Spotifyapk56"
        
        # Try to get chat info
        try:
            chat = await client.get_chat(channel_username)
            chat_id = chat.id
            chat_title = chat.title or chat.username or channel_username
            
            # Check for last indexed message to resume from there
            last_indexed = db.get_stored_last_indexed_message_id(chat_id)
            
            if last_indexed >= 18097631:
                # Resume from last indexed position
                start_message_id = last_indexed + 1
                logger.info(f"Resuming indexing from message {start_message_id} (last indexed: {last_indexed})")
                await message.reply(f"🔄 Resuming indexing from message {start_message_id}\n"
                                  f"Last indexed message: {last_indexed}")
                
                # Find the current latest message for target
                try:
                    recent_messages = await client.get_chat_history(chat_id, limit=1)
                    if recent_messages:
                        target_message_id = recent_messages[0].id
                        logger.info(f"Target message ID (latest): {target_message_id}")
                    else:
                        target_message_id = start_message_id + 1000  # Fallback
                except Exception as e:
                    logger.warning(f"Could not get latest message, using fallback: {e}")
                    target_message_id = start_message_id + 1000
                    
            else:
                # Start from the specified first message ID: 18149251
                start_message_id = 18097631
                logger.info(f"Starting indexing from first message: {start_message_id}")
                await message.reply(f"🚀 Starting indexing from first message: {start_message_id}")
                
                # Find the current latest message for target
                try:
                    recent_messages = await client.get_chat_history(chat_id, limit=1)
                    if recent_messages:
                        target_message_id = recent_messages[0].id
                        logger.info(f"Target message ID (latest): {target_message_id}")
                        await message.reply(f"📍 Latest message found: {target_message_id}")
                    else:
                        target_message_id = start_message_id + 10000  # Large range fallback
                        logger.warning("Could not get latest message, using large range")
                except Exception as e:
                    logger.warning(f"Could not get latest message: {e}")
                    target_message_id = start_message_id + 10000
            
            logger.info(f"Auto-starting indexing from {chat_title} (ID: {chat_id})")
            logger.info(f"Range: {start_message_id} to {target_message_id}")
            
            await start_indexing_process(client, message, chat_id, start_message_id, chat_title, target_message_id)
            
            # After completing initial indexing, start continuous monitoring
            asyncio.create_task(start_continuous_monitoring(client, chat_id, target_message_id, chat_title))
            
        except Exception as e:
            logger.error(f"Failed to auto-start indexing: {e}")
            await message.reply(f"❌ Failed to auto-start indexing from Spotifyapk56: {str(e)}")
            
    except Exception as e:
        logger.error(f"Error in auto_start_spotifyapk_indexing: {e}")

async def find_current_first_message(client: Client, chat_id: int) -> int:
    """Find the current first available message in the channel"""
    try:
        logger.info(f"Searching for first available message in chat {chat_id}")
        
        # Start searching from message ID 1 and work upwards
        test_message_id = 1
        max_attempts = 1000  # Don't search too long
        
        while test_message_id <= max_attempts:
            try:
                # Try to get the message
                msg = await client.get_messages(chat_id, test_message_id)
                if msg and not msg.empty:
                    logger.info(f"Found first available message: {test_message_id}")
                    return test_message_id
                    
            except Exception as e:
                # Message doesn't exist, try next
                if "MESSAGE_ID_INVALID" in str(e):
                    test_message_id += 1
                    continue
                else:
                    logger.debug(f"Error checking message {test_message_id}: {e}")
                    test_message_id += 1
                    continue
            
            test_message_id += 1
            
            # Add small delay to avoid rate limits during search
            if test_message_id % 10 == 0:
                await asyncio.sleep(0.1)
        
        logger.warning(f"Could not find first message after checking {max_attempts} messages")
        return None
        
    except Exception as e:
        logger.error(f"Error finding first message: {e}")
        return None

async def start_continuous_monitoring(client: Client, chat_id: int, last_processed_id: int, chat_title: str):
    """Start continuous monitoring for new messages after initial indexing completes"""
    global indexing_process
    
    # Wait for initial indexing to complete
    while indexing_process["active"]:
        await asyncio.sleep(5)
        
    logger.info(f"Starting continuous monitoring for {chat_title} from message ID {last_processed_id + 1}")
    
    # Monitor for new messages every 30 seconds
    while True:
        try:
            await asyncio.sleep(30)  # Check every 30 seconds
            
            # Get the latest message ID
            try:
                # Get recent messages to find the latest message ID
                messages = await client.get_chat_history(chat_id, limit=1)
                if messages:
                    latest_msg_id = messages[0].id
                    
                    # If there are new messages, process them
                    if latest_msg_id > last_processed_id:
                        logger.info(f"Found new messages in {chat_title}: {last_processed_id + 1} to {latest_msg_id}")
                        
                        # Process new messages
                        for msg_id in range(last_processed_id + 1, latest_msg_id + 1):
                            try:
                                msg = await client.get_messages(chat_id, msg_id)
                                if msg and (msg.audio or msg.video or msg.document or msg.photo):
                                    await handle_media_message(client, msg)
                                    logger.info(f"Auto-processed new media message {msg_id}")
                                    
                                # Small delay to prevent rate limiting
                                await asyncio.sleep(0.5)
                                
                            except Exception as e:
                                logger.debug(f"Could not process message {msg_id}: {e}")
                                continue
                        
                        # Update the last processed ID
                        last_processed_id = latest_msg_id
                        
            except Exception as e:
                logger.error(f"Error checking for new messages: {e}")
                continue
                
        except Exception as e:
            logger.error(f"Error in continuous monitoring: {e}")
            # Wait longer before retrying on error
            await asyncio.sleep(60)

def setup_handlers(app: Client):
    """Setup all message handlers"""
    
    # Media message handlers (handle media from all chats for indexing)
    app.on_message(filters.audio & ~filters.bot)(handle_media_message)
    app.on_message(filters.video & ~filters.bot)(handle_media_message)
    app.on_message(filters.document & ~filters.bot)(handle_media_message)
    app.on_message(filters.photo & ~filters.bot)(handle_media_message)
    
    # Command handlers
    app.on_message(filters.command("start"))(handle_start_command)
    app.on_message(filters.command("send"))(handle_send_command)
    app.on_message(filters.command("sendid"))(handle_sendid_command)
    app.on_message(filters.command("stats"))(handle_stats_command)
    app.on_message(filters.command("stop_index"))(handle_stop_index_command)
    app.on_message(filters.command("cancel"))(handle_stop_index_command)  # /cancel works same as /stop_index
    app.on_message(filters.command("db"))(handle_db_command)
    
    # Message link handler
    app.on_message(filters.text & filters.regex(r"https://t\.me/[^/]+/\d+") & ~filters.bot)(handle_message_link)
    
    # Forwarded message handler for automatic indexing
    app.on_message(filters.forwarded & ~filters.bot)(handle_forwarded_message)
    
    logger.info("All handlers registered successfully")
